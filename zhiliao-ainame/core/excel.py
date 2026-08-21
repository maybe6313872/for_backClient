"""
Excel 文件处理模块

本模块提供了通用的 Excel 文件读取和写入功能。
不包含业务逻辑，只负责 Excel 文件的解析和生成。
"""

from io import BytesIO
from typing import List, Dict, Tuple, Optional, Any
from datetime import datetime
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def create_excel_file(
    headers: List[str],
    data_rows: List[List[Any]],
    sheet_title: str = "Sheet1",
    column_widths: Optional[Dict[str, int]] = None,
    header_style: Optional[Dict[str, Any]] = None
) -> BytesIO:
    """
    创建 Excel 文件
    
    通用的 Excel 文件生成函数，不包含业务逻辑。
    
    Args:
        headers (List[str]): 表头列表
        data_rows (List[List[Any]]): 数据行列表，每个元素是一行的数据列表
        sheet_title (str): 工作表标题，默认为 "Sheet1"
        column_widths (Optional[Dict[str, int]]): 列宽字典，键为列字母（如 'A', 'B'），值为宽度
        header_style (Optional[Dict[str, Any]]): 表头样式配置，可选键：
            - font_color: 字体颜色（如 "FFFFFF"）
            - fill_color: 填充颜色（如 "366092"）
            - bold: 是否加粗，默认 True
            - alignment: 对齐方式，默认 "center"
        
    Returns:
        BytesIO: Excel 文件的二进制流
    """
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # 设置表头
    ws.append(headers)
    
    # 设置表头样式（如果提供了样式配置）
    if header_style:
        font_color = header_style.get('font_color', 'FFFFFF')
        fill_color = header_style.get('fill_color', '366092')
        bold = header_style.get('bold', True)
        alignment = header_style.get('alignment', 'center')
        
        header_font = Font(bold=bold, color=font_color)
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        header_alignment = Alignment(horizontal=alignment, vertical=alignment)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    # 填充数据
    for row in data_rows:
        ws.append(row)
    
    # 调整列宽（如果提供了列宽配置）
    if column_widths:
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    # 将工作簿保存到内存
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def encode_filename_for_download(filename: str) -> str:
    """
    对文件名进行 URL 编码，以支持中文文件名
    
    使用 RFC 5987 格式：filename*=UTF-8'' 用于 UTF-8 编码的文件名
    
    Args:
        filename (str): 原始文件名（可能包含中文）
        
    Returns:
        str: URL 编码后的文件名
    """
    return quote(filename, safe='')


def parse_excel(
    file_content: bytes,
    start_row: int = 1,
    header_row: int = 1,
    data_start_row: int = 2
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    解析 Excel 文件，返回表头和数据行
    
    通用的 Excel 解析函数，不包含业务逻辑，只负责读取和解析数据。
    
    Args:
        file_content (bytes): Excel 文件的二进制内容
        start_row (int): 开始读取的行号，默认为 1
        header_row (int): 表头所在的行号，默认为 1
        data_start_row (int): 数据开始的行号，默认为 2（表头下一行）
        
    Returns:
        Tuple[List[str], List[Dict[str, Any]]]:
            - 表头列表：表头字段名列表
            - 数据行列表：每行是一个字典，键为列索引（1-based），值为单元格值
            
    Raises:
        ValueError: 当文件无法读取时
    """
    excel_buffer = BytesIO(file_content)
    
    # 使用 openpyxl 打开 Excel 文件
    try:
        wb = openpyxl.load_workbook(excel_buffer, data_only=True)
    except Exception as e:
        raise ValueError(f"无法读取Excel文件: {str(e)}")
    
    # 获取第一个工作表
    ws = wb.active
    
    # 读取表头
    header_cells = list(ws[header_row])
    headers = [cell.value if cell.value is not None else f"列{idx+1}" for idx, cell in enumerate(header_cells)]
    
    # 解析数据行
    data_rows = []
    for row_num in range(data_start_row, ws.max_row + 1):
        row = ws[row_num]
        row_data = {
            'row_num': row_num,
            'cells': {}
        }
        
        # 将每列的值存储到字典中，键为列索引（1-based）
        for col_idx, cell in enumerate(row, start=1):
            row_data['cells'][col_idx] = cell.value
        
        data_rows.append(row_data)
    
    return headers, data_rows


def find_column_indices(headers: List[str], column_names: Dict[str, List[str]]) -> Dict[str, int]:
    """
    根据表头查找列的索引位置
    
    在表头中查找匹配的列名，返回列索引字典（1-based）。
    
    Args:
        headers (List[str]): 表头列表
        column_names (Dict[str, List[str]]): 要查找的列名映射，键为自定义名称，值为可能的表头名称列表
            例如：{'username': ['用户名', 'username', '用户']}
        
    Returns:
        Dict[str, int]: 列索引字典，键为自定义名称，值为列索引（1-based），未找到的列值为 None
    """
    result = {}
    
    for custom_name, possible_names in column_names.items():
        result[custom_name] = None
        for idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            header_lower = str(header).strip().lower()
            if header_lower in [name.lower() for name in possible_names]:
                result[custom_name] = idx
                break
    
    return result
